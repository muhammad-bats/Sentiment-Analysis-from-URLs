import PyPDF2 as PDF
import vaderSentiment
import requests
import openpyxl
import nltk
import numpy as np
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from nltk.tokenize import sent_tokenize

totalpos = 0
count = 0

def get_file(url):
   fname = input("Enter file name you want ")
   response = requests.get(url)
   if response.status_code == 200:
      print("File downloaded successfully")
      with open(fname, "wb") as file:
         file.write(response.content)
   else: 
      print("Could not download file: ", response.status_code)
      fname = "-1"
   return fname

def extract_text(pdf_file):
  
  with open(pdf_file, 'rb') as fh:
    fh = PDF.PdfReader(fh)
    num_pages = len(fh.pages)
    text = ""
    for i in range(num_pages):
      page = fh.pages[i]
      text += page.extract_text()
  return text  

def tokenize(text):
    sentences = sent_tokenize(text)
    return sentences


def sentiment_analyze(sentences):
    result = dict()
    sentpos = list()
    sentneg = list()
    sentneu = list()
    sentcom = list()
    analyzer = SentimentIntensityAnalyzer()
    for sentence in sentences:
       sentresult = analyzer.polarity_scores(sentence)
       sentpos.append(sentresult['pos'])
       sentneg.append(sentresult['neg'])
       sentneu.append(sentresult['neu'])
       sentcom.append(sentresult['compound'])
       
    avgpos = np.mean(sentpos)
    avgneg = np.mean(sentneg)
    avgneu = np.mean(sentneu)
    avgcom = np.mean(sentcom)
    result['pos'] = avgpos
    result['neg'] = avgneg
    result['neu'] = avgneu
    result['compound'] = avgcom
    return result

def analysis(text, totalpos):
   data = list()
   result = sentiment_analyze(text)
   poscore = result['pos']
   negscore = result['neg']
   neuscore = result['neu']
   totalpos += poscore
   print(result)
   if result['compound'] > 0:
            category = 'POSITIVE'
            value = result['compound']
            print("Text has sentiment: ", category, "value: ", value)
   elif result['compound'] < 0:
        category = 'NEGATIVE'
        value = result['compound']
        print("Text has sentiment: ", category, "value: ", value)
   else:
      category = 'NEUTRAL'
      value = result['compound']
      print("Text has sentiment: ", category, "value: ", value)
    
   data.append(url)
   data.append(category)
   data.append(poscore)
   data.append(negscore)
   data.append(neuscore)
   data.append(value)
   datalst.append(data)
   print(datalst)
   return totalpos

def write_excel(data, filename):
  try:
    workbook = openpyxl.load_workbook(filename)
  except FileNotFoundError:
    workbook = openpyxl.Workbook()
  sheet = workbook.active
  
  last_row = sheet.max_row + 1
  for row_num, row_data in enumerate(data, start=last_row):
    for col_num, value in enumerate(row_data, start=1):
      cell = sheet.cell(row=row_num, column=col_num)
      cell.value = value

  workbook.save(filename)
  print("Data written to file")

# press 'y' to continue loop
totalpos = 0
count = 0
ok = True
while ok:
    datalst = list()
    url = input("Enter URL: ")
    fname = get_file(url)
    if fname != "-1":
        text = extract_text(fname)
        sentences = tokenize(text)
        totalpos = analysis(sentences, totalpos)
        fexcel = 'finalexcel.xlsx'
        write_excel(datalst, fexcel)
        print("done")
        count += 1
        cont = input("continue? ")
        if cont != "y":
            ok = False 
    else: break
pospercent = (totalpos/count)*100
print("Average Positivity Percentage is: ", pospercent)